from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from accounts.models import StudentProfile, ContactMessage
from events.models import Event, EventRegistration
from payments.models import Payment
from events.models import Event, EventRegistration
from payments.models import Payment 
from .models import AdminProfile
from .forms import AdminProfileForm
from accounts.forms import StudentProfileForm
from django.http import HttpResponse
import csv
import openpyxl  
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils.timezone import localtime
from django.db.models import Q
from django.core.paginator import Paginator
from events.models import Event
from .forms import EventForm 
from django.contrib import messages
from datetime import datetime
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render, redirect
from django.contrib.admin.views.decorators import staff_member_required
from events.models import Announcement
from .forms import AnnouncementForm


# from adminpanel.models import EventRegistration





# ------------------- PERMISSION CHECK -------------------
def admin_required(view_func):
    return user_passes_test(lambda u: u.is_staff or u.is_superuser)(view_func)


# ------------------- ADMIN DASHBOARD -------------------
@login_required(login_url='admin_login')
@admin_required
def admin_dashboard(request):
    students_count = StudentProfile.objects.count()
    events_count = Event.objects.count()
    payments_count = Payment.objects.count()
    registrations_count = EventRegistration.objects.count()
    feedback_count = ContactMessage.objects.count()

    students = StudentProfile.objects.all()[:5]  # latest students
    events = Event.objects.all().order_by("-date")[:5]  # recent events
    payments = Payment.objects.select_related('student__user').order_by("-created_at")[:5]  
    feedbacks = ContactMessage.objects.all().order_by("-created_at")[:5]

    context = {
        "students_count": students_count,
        "events_count": events_count,
        "payments_count": payments_count,
        "registrations_count": registrations_count,
        "feedback_count": feedback_count,
        "students": students,
        "events": events,
        "payments": payments,
        "feedbacks": feedbacks,
    }
    return render(request, "adminpanel/admin_dashboard.html", context)




# ------------------- ADMIN PROFILE -------------------
@login_required(login_url='admin_login')
@admin_required
def admin_profile(request):
    profile, created = AdminProfile.objects.get_or_create(user=request.user)

    if request.method == "POST":
        form = AdminProfileForm(request.POST, request.FILES, instance=profile)
        if form.is_valid():
            form.save()
            messages.success(request, "Profile updated successfully!")
            return redirect("admin_profile")
    else:
        form = AdminProfileForm(instance=profile)

    return render(request, "adminpanel/profile.html", {"form": form})


# ------------------- UPGRADE OPTIONS -------------------
# @login_required
# @admin_required
# def upgrade_option_list(request):
#     upgrades = UpgradeOption.objects.all()
#     return render(request, "adminpanel/upgrade_list.html", {"upgrades": upgrades})


# @login_required
# @admin_required
# def upgrade_option_detail(request, pk):
#     upgrade = get_object_or_404(UpgradeOption, pk=pk)
#     return render(request, "adminpanel/upgrade_detail.html", {"upgrade": upgrade})


# @login_required
# @admin_required
# def upgrade_option_create(request):
#     if request.method == "POST":
#         form = UpgradeOptionForm(request.POST)
#         if form.is_valid():
#             form.save()
#             messages.success(request, "Upgrade option created successfully!")
#             return redirect("upgrade_option_list")
#     else:
#         form = UpgradeOptionForm()
#     return render(request, "adminpanel/upgrade_form.html", {"form": form})


# @login_required
# @admin_required
# def upgrade_option_edit(request, pk):
#     upgrade = get_object_or_404(UpgradeOption, pk=pk)
#     if request.method == "POST":
#         form = UpgradeOptionForm(request.POST, instance=upgrade)
#         if form.is_valid():
#             form.save()
#             messages.success(request, "Upgrade option updated successfully!")
#             return redirect("upgrade_option_list")
#     else:
#         form = UpgradeOptionForm(instance=upgrade)
#     return render(request, "adminpanel/upgrade_form.html", {"form": form})


# @login_required
# @admin_required
# def upgrade_option_delete(request, pk):
#     upgrade = get_object_or_404(UpgradeOption, pk=pk)
#     if request.method == "POST":
#         upgrade.delete()
#         messages.success(request, "Upgrade option deleted successfully!")
#         return redirect("upgrade_option_list")
#     return render(request, "adminpanel/upgrade_detail.html", {"upgrade": upgrade})


# ------------------- STUDENT MANAGEMENT -------------------
@login_required(login_url='admin_login')
@admin_required
def student_list(request):
    students = StudentProfile.objects.all()
    return render(request, "adminpanel/students_list.html", {"students": students})


@login_required(login_url='admin_login')
@admin_required
def student_detail(request, pk):
    student = get_object_or_404(StudentProfile, pk=pk)
    return render(request, "adminpanel/student_detail.html", {"student": student})


# ------------------- EVENT MANAGEMENT -------------------
@login_required(login_url='admin_login')
@admin_required
def admin_event_list(request):
    query = request.GET.get("q", "").strip()
    events_qs = Event.objects.all().order_by("-date")

    if query:
        events_qs = events_qs.filter(
            Q(title__icontains=query) |
            Q(description__icontains=query) |
            Q(location__icontains=query)
        )

    paginator = Paginator(events_qs, 10)  # change page size here if you want
    page_number = request.GET.get("page")
    events = paginator.get_page(page_number)

    context = {
        "events": events,
        "query": query,
    }
    return render(request, "adminpanel/events_list.html", context)





# ------------------- PAYMENT MANAGEMENT -------------------
@login_required(login_url='admin_login')
@admin_required
def payment_list(request):
    payments = Payment.objects.all()
    return render(request, "adminpanel/payments_list.html", {"payments": payments})


@login_required(login_url='admin_login')
@admin_required
def payment_detail(request, pk):
    payment = get_object_or_404(Payment, pk=pk)
    return render(request, "adminpanel/payment_detail.html", {"payment": payment})


# ------------------- FEEDBACK MANAGEMENT -------------------
@login_required(login_url='admin_login')
@admin_required
def feedback_list(request):
    feedbacks = ContactMessage.objects.all().order_by("-created_at")
    return render(request, "adminpanel/feedback_list.html", {"feedbacks": feedbacks})


@login_required(login_url='admin_login')
@admin_required
def feedback_delete(request, pk):
    feedback = get_object_or_404(ContactMessage, pk=pk)
    if request.method == "POST":
        feedback.delete()
        messages.success(request, "Feedback deleted successfully!")
        return redirect("feedback_list")
    return render(request, "adminpanel/feedback_detail.html", {"feedback": feedback})


@login_required(login_url='admin_login')
@admin_required
def student_edit(request, pk):
    student = get_object_or_404(StudentProfile, pk=pk)
    
    if request.method == "POST":
        form = StudentProfileForm(request.POST, request.FILES, instance=student)
        if form.is_valid():
            form.save()
            messages.success(request, "Student profile updated successfully!")
            return redirect("student_detail", pk=student.pk)
    else:
        form = StudentProfileForm(instance=student)

    return render(request, "adminpanel/student_form.html", {"form": form, "student": student})


@login_required(login_url='admin_login')
@admin_required
def student_delete(request, pk):
    student = get_object_or_404(StudentProfile, pk=pk)
    
    if request.method == "POST":
        student.delete()
        messages.success(request, "Student deleted successfully!")
        return redirect("student_list")
    
    return render(request, "adminpanel/student_confirm_delete.html", {"student": student})


@login_required(login_url='admin_login')
@admin_required
def event_edit(request, pk):
    event = get_object_or_404(Event, pk=pk)
    from adminpanel.forms import EventForm  # import here to be safe

    if request.method == "POST":
        form = EventForm(request.POST, request.FILES, instance=event)
        if form.is_valid():
            form.save()
            messages.success(request, "Event updated successfully!")
            # keep existing detail redirect; change if you want admin listing instead
            return redirect("event_detail", pk=event.pk)
    else:
        form = EventForm(instance=event)

    return render(request, "adminpanel/event_form.html", {"form": form, "event": event})


@login_required(login_url='admin_login')
@admin_required
def event_delete(request, pk):
    event = get_object_or_404(Event, pk=pk)
    if request.method == "POST":
        event.delete()
        messages.success(request, "Event deleted successfully!")
        return redirect("admin_event_list")
    return render(request, "adminpanel/event_confirm_delete.html", {"event": event})


@login_required(login_url='admin_login')
@admin_required
def payment_edit(request, pk):
    payment = get_object_or_404(Payment, pk=pk)
    if request.method == "POST":
        from payments.forms import PaymentForm  # ensure you have this form
        form = PaymentForm(request.POST, instance=payment)
        if form.is_valid():
            form.save()
            messages.success(request, "Payment updated successfully!")
            return redirect("payment_detail", pk=payment.pk)
    else:
        from payments.forms import PaymentForm
        form = PaymentForm(instance=payment)

    return render(request, "adminpanel/payment_form.html", {"form": form})


@login_required(login_url='admin_login')
@admin_required
def payment_delete(request, pk):
    payment = get_object_or_404(Payment, pk=pk)
    if request.method == "POST":
        payment.delete()
        messages.success(request, "Payment deleted successfully!")
        return redirect("payment_list")

    return render(request, "adminpanel/payment_confirm_delete.html", {"payment": payment})





# @login_required
# @admin_required
# def upgrade_create(request):
#     if request.method == "POST":
#         form = UpgradeOptionForm(request.POST)
#         if form.is_valid():
#             form.save()
#             messages.success(request, "New upgrade option created successfully!")
#             return redirect("upgrade_option_list")
#     else:
#         form = UpgradeOptionForm()

#     return render(request, "adminpanel/upgrade_form.html", {"form": form})




# Admin: List all event registrations
def admin_event_registrations(request):
    registrations = EventRegistration.objects.select_related("student", "event").all()
    return render(request, "adminpanel/event_registrations.html", {"registrations": registrations})

# Admin: List all events
from django.core.paginator import Paginator
from events.models import Event

@login_required(login_url='admin_login')
def admin_events(request):
    query = request.GET.get("q", "")
    events = Event.objects.all().order_by("-date")

    # Search functionality
    if query:
        events = events.filter(title__icontains=query)

    # Pagination
    paginator = Paginator(events, 10)  # 10 events per page
    page_number = request.GET.get("page")
    events_page = paginator.get_page(page_number)

    return render(request, "adminpanel/events_list.html", {
        "events": events_page,
        "query": query,
    })


def add_event(request):
    if request.method == 'POST':
        form = EventForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('admin_event_list')
    else:
        form = EventForm()
    return render(request, 'adminpanel/event_form.html', {'form': form})


# @login_required
# @admin_required
# def admin_edit_event(request, pk):
#     event = get_object_or_404(Event, pk=pk)
#     from adminpanel.forms import EventForm
#     if request.method == "POST":
#         form = EventForm(request.POST, instance=event)
#         if form.is_valid():
#             form.save()
#             messages.success(request, "Event updated successfully!")
#             return redirect("admin_events")
#     else:
#         form = EventForm(instance=event)
#     return render(request, "adminpanel/event_form.html", {"form": form, "event": event})



# Payments (read-only)
def admin_payments(request):
    payments = Payment.objects.select_related("student").all()
    return render(request, "adminpanel/payments.html", {"payments": payments})


# Export Payments CSV
def export_payments_csv(request):
    payments = Payment.objects.select_related("student").all()

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="payments.csv"'

    writer = csv.writer(response)
    writer.writerow(["ID", "Student", "Email", "Amount", "Status", "Date"])
    for p in payments:
        writer.writerow([
            p.id,
            p.student.user.username,
            p.student.user.email,
            p.amount,
            p.status,
            p.created_at.strftime("%Y-%m-%d %H:%M"),
        ])
    return response



def export_registrations_excel(request):
    registrations = EventRegistration.objects.select_related("student__user", "event")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registrations"

    ws.append(["Student", "Email", "Event", "Registered At"])

    for reg in registrations:
        formatted_time = localtime(reg.registered_at).strftime("%b %d, %Y %I:%M %p")
        ws.append([
            reg.student.user.username,
            reg.student.user.email,
            reg.event.title,
            formatted_time
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="event_registrations.xlsx"'
    wb.save(response)

    return response




@login_required(login_url='admin_login')
@admin_required
def event_create(request):
    from adminpanel.forms import EventForm

    if request.method == "POST":
        form = EventForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "New event created successfully!")
            return redirect("admin_event_list")  
    else:
        form = EventForm()
    return render(request, "adminpanel/event_form.html", {"form": form})



@login_required(login_url='admin_login')
@admin_required
def export_event_registrations_csv(request, pk):
    event = get_object_or_404(Event, pk=pk)

    registrations = EventRegistration.objects.filter(event=event)

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="event_{event.id}_registrations.csv"'

    writer = csv.writer(response)
    writer.writerow(["Student", "Email", "Registered At"])

    for reg in registrations:
        writer.writerow([
            reg.student.user.get_full_name() or reg.student.user.username,
            reg.student.user.email,
            reg.registered_at.strftime("%Y-%m-%d %H:%M"),
        ])

    return response


@login_required(login_url='admin_login')
@admin_required
def export_event_registrations_excel(request, pk):
    event = get_object_or_404(Event, pk=pk)

    registrations = EventRegistration.objects.filter(event=event)

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{event.title[:20]} Registrations"

    # Header row
    headers = ["Student", "Email", "Registered At"]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = header

    # Data rows
    for row_num, reg in enumerate(registrations, 2):
        ws[f"A{row_num}"] = reg.student.user.get_full_name() or reg.student.user.username
        ws[f"B{row_num}"] = reg.student.user.email
        ws[f"C{row_num}"] = reg.registered_at.strftime("%Y-%m-%d %I:%M %p")

    # Prepare response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="event_{event.id}_registrations.xlsx"'
    wb.save(response)

    return response

@login_required(login_url='admin_login')
@admin_required
def admin_event_detail(request, pk):
    event = get_object_or_404(Event, pk=pk)
    registrations = EventRegistration.objects.filter(event=event)

    return render(request, "adminpanel/event_detail.html", {
        "event": event,
        "registrations": registrations,
    })




@login_required(login_url='admin_login')
@admin_required
def admin_edit_event(request, pk):
    event = get_object_or_404(Event, pk=pk)
    from adminpanel.forms import EventForm  

    if request.method == "POST":
        form = EventForm(request.POST, request.FILES, instance=event)
        if form.is_valid():
            form.save()
            messages.success(request, "Event updated successfully!")
            return redirect("event_detail", pk=event.pk)
    else:
        form = EventForm(instance=event)

    return render(request, "adminpanel/event_form.html", {"form": form, "event": event})









def admin_login(request):
    # if request.user.is_authenticated:
    #     if hasattr(request.user, 'adminprofile'):
    #         return redirect("admin_dashboard")
    #     else:
    #         return redirect("home")  

    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(request, username=username, password=password)

        if user is not None and hasattr(user, 'adminprofile'):
            login(request, user)
            messages.success(request, f"Welcome back, {user.username}!")
            return redirect("admin_dashboard")
        else:
            messages.error(request, "Invalid admin credentials.")

    return render(request, "adminpanel/admin_login.html")


@login_required(login_url='admin_login')
def admin_logout(request):
    logout(request)
    messages.success(request, "You have been logged out successfully.")
    return redirect("admin_login")




@staff_member_required
def announcements_list(request):
    qs = Announcement.objects.order_by("-is_pinned", "-created_at")
    return render(request, "adminpanel/announcements_list.html", {"announcements": qs})

@staff_member_required
def announcement_add(request):
    if request.method == "POST":
        form = AnnouncementForm(request.POST)
        if form.is_valid():
            ann = form.save(commit=False)
            ann.created_by = request.user
            ann.save()
            messages.success(request, "Announcement created.")
            return redirect("adminpanel_announcements")
    else:
        form = AnnouncementForm()
    return render(request, "adminpanel/announcement_form.html", {"form": form, "action": "Add"})

@staff_member_required
def announcement_edit(request, pk):
    ann = get_object_or_404(Announcement, pk=pk)
    if request.method == "POST":
        form = AnnouncementForm(request.POST, instance=ann)
        if form.is_valid():
            form.save()
            messages.success(request, "Announcement updated.")
            return redirect("adminpanel_announcements")
    else:
        form = AnnouncementForm(instance=ann)
    return render(request, "adminpanel/announcement_form.html", {"form": form, "action": "Edit", "announcement": ann})

@staff_member_required
def announcement_delete(request, pk):
    ann = get_object_or_404(Announcement, pk=pk)
    if request.method == "POST":
        ann.delete()
        messages.success(request, "Announcement deleted.")
        return redirect("adminpanel_announcements")
    return render(request, "adminpanel/announcement_confirm_delete.html", {"announcement": ann})
